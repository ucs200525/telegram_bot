import platform
import pandas as pd
import requests
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import Application, CommandHandler, CallbackContext, ConversationHandler, MessageHandler, filters
import pytz
from openpyxl import load_workbook
from opencage.geocoder import OpenCageGeocode
import win32com.client
from PIL import ImageGrab
import logging
import shutil

# Set up logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Define conversation state
LOCATION = 0

# Function to get sunrise and sunset times
def get_sun_times(lat, lng, local_tz):
    print("Fetching sunrise and sunset times")
    today = datetime.now().strftime('%Y-%m-%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

    url = f'https://api.sunrise-sunset.org/json?lat={lat}&lng={lng}&formatted=0&date='

    today_response = requests.get(url + today).json()
    tomorrow_response = requests.get(url + tomorrow).json()

    sunrise_today_utc = today_response['results']['sunrise']
    sunset_today_utc = today_response['results']['sunset']
    sunrise_tomorrow_utc = tomorrow_response['results']['sunrise']

    # Convert to local time
    ist = pytz.timezone(local_tz)
    sunrise_today = pd.to_datetime(sunrise_today_utc).tz_convert(ist)
    sunset_today = pd.to_datetime(sunset_today_utc).tz_convert(ist)
    sunrise_tomorrow = pd.to_datetime(sunrise_tomorrow_utc).tz_convert(ist)

    return sunrise_today, sunset_today, sunrise_tomorrow

# Function to update Excel file
def update_excel(file_path, sunrise_today, sunset_today, sunrise_tomorrow):
    print("Updating the Excel file")
    # Copy the file to avoid permission issues
    temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
    shutil.copy(file_path, temp_file_path)

    # Load the workbook and select the active worksheet
    wb = load_workbook(temp_file_path)
    ws = wb.active
    
    # Update the specific cells
    ws['O4'] = sunrise_today.strftime('%H:%M:%S')
    ws['O5'] = sunset_today.strftime('%H:%M:%S')
    ws['O6'] = sunrise_tomorrow.strftime('%H:%M:%S')

    # Save the workbook
    wb.save(temp_file_path)
    print("Excel file updated")

    return temp_file_path

# Function to save Excel range as image
def save_excel_range_as_image(file_path, save_path):
    print("Saving Excel range as image")
    if platform.system() == 'Windows':
        # Initialize Excel application
        excel = win32com.client.Dispatch('Excel.Application')
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(Filename=file_path)
        
        # Select worksheet (assuming the first sheet)
        ws = wb.Worksheets(1)
        
        # Define the range to copy (B-L, 1-32)
        start_row = 1
        start_col = 2  # Column B
        end_row = 32
        end_col = 12  # Column L
        
        # Copy range as picture
        ws.Range(ws.Cells(start_row, start_col), ws.Cells(end_row, end_col)).CopyPicture(Format=2)
        
        # Create a new workbook to paste the copied picture
        wb_new = excel.Workbooks.Add()
        ws_new = wb_new.ActiveSheet
        ws_new.Paste()

        # Check for shapes
        shapes = ws_new.Shapes
        shape_names = [shape.Name for shape in shapes]
        print(f"Shapes in worksheet: {shape_names}")

        # Copy the picture as a shape
        if 'Picture 1' in shape_names:
            ws_new.Shapes('Picture 1').Copy()
            
            # Grab the image from clipboard and save
            img = ImageGrab.grabclipboard()
            if img:
                img.save(save_path)
                print(f'Image saved successfully at: {save_path}')
            else:
                print('Failed to capture image from clipboard.')
        else:
            print("Shape 'Picture 1' not found.")
        
        # Close workbook without saving changes
        wb.Close(SaveChanges=False)
        wb_new.Close(SaveChanges=False)
        
        # Quit Excel application
        excel.Quit()
        print("Excel application closed")
    else:
        print('Error: This functionality is only supported on Windows.')

# Command handler to start the conversation
async def send_table_start(update: Update, context: CallbackContext):
    await update.message.reply_text("Please enter your location (Vijayawada):")
    return LOCATION

# Function to handle location input and send the table with image
async def receive_location(update: Update, context: CallbackContext):
    location = update.message.text
    print(f"Received location: {location}")

    # Use OpenCage Geocoder to get coordinates
    geocoder = OpenCageGeocode(context.bot_data['opencage_api_key'])
    result = geocoder.geocode(location)

    if result and len(result):
        latitude = result[0]['geometry']['lat']
        longitude = result[0]['geometry']['lng']
        
        local_tz = 'Asia/Kolkata'  # Assuming Indian Standard Time (IST)
        
        # Get sun times
        sunrise_today, sunset_today, sunrise_tomorrow = get_sun_times(latitude, longitude, local_tz)

        # Update the Excel file
        file_path = context.bot_data['excel_file_path']
        try:
            updated_file_path = update_excel(file_path, sunrise_today, sunset_today, sunrise_tomorrow)

            # Save Excel range as image
            save_image_path = context.bot_data['image_save_path']
            save_excel_range_as_image(updated_file_path, save_image_path)

            # Send the saved image
            print("Sending the saved image")
            await update.message.reply_photo(photo=open(save_image_path, 'rb'))
        except Exception as e:
            logging.error(f"Error updating Excel file: {e}")
            await update.message.reply_text(f"An error occurred while updating the Excel file: {e}")

        # End the conversation
        return ConversationHandler.END
    else:
        await update.message.reply_text("Sorry, I couldn't find coordinates for that location. Please try again.")
        return LOCATION

async def help_command_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("This is the help message.")

async def main_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("You sent a text message.")

def main():
    # Prompt user to enter tokens and paths
    opencage_api_key = '699522e909454a09b82d1c728fc79925'
    excel_file_path = '/usr/src/app/Copy Completed_eng_Bharghava_Siddhanta_Panchangam.xlsx'
    image_save_path =  '/usr/src/app/imageTo.png'
    
        
    #'C:\Users\sree\telegram_bot\Copy Completed_eng_Bharghava_Siddhanta_Panchangam.xlsx'*/
    #r'C:\\Users\\sree\\telegram_bot\\imageTo.png'
   
    bot_token ='7274941037:AAHIWiU5yvfIzo7eJWPu9S5CeJIid6ATEyM'


    # Create the Application instance
    application = Application.builder().token(bot_token).build()

    # Save tokens and paths in bot_data
    application.bot_data['opencage_api_key'] = opencage_api_key
    application.bot_data['excel_file_path'] = excel_file_path
    application.bot_data['image_save_path'] = image_save_path

    # Create the conversation handler
    conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('panchangam', send_table_start)],
        states={
            LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_location)],
        },
        fallbacks=[]
    )

    # Add conversation handler
    application.add_handler(conversation_handler)

    # Add other handlers
    application.add_handler(CommandHandler("help", help_command_handler))
    application.add_handler(MessageHandler(filters.TEXT, main_handler))

    # Run the bot
    application.run_polling()

if __name__ == '__main__':
    main()
